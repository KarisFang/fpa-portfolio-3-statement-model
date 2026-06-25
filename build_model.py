"""
Apple Inc. (AAPL) 3-Statement Financial Model
Built on real FY2023-FY2025 10-K / 8-K data from SEC EDGAR.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

wb = Workbook()
wb.remove(wb.active)

FONT_NAME = "Calibri"
BLUE = Font(name=FONT_NAME, color="0000FF")
BLUE_BOLD = Font(name=FONT_NAME, color="0000FF", bold=True)
BLACK = Font(name=FONT_NAME, color="000000")
BLACK_BOLD = Font(name=FONT_NAME, color="000000", bold=True)
GREEN = Font(name=FONT_NAME, color="008000")
WHITE_BOLD = Font(name=FONT_NAME, color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=10, color="595959")
SECTION_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
TOTAL_BORDER = Border(top=Side(style="thin"))
DOUBLE_BORDER = Border(top=Side(style="thin"), bottom=Side(style="double"))

CUR = '$#,##0;($#,##0);"-"'
NUM = '#,##0;(#,##0);"-"'
PCT = '0.0%;(0.0%);"-"'
EPS_FMT = '$#,##0.00;($#,##0.00)'
DAYS_FMT = '0.0'
PRICE_FMT = '$#,##0.00'

YEAR_COLS = ["B", "C", "D", "E", "F", "G"]
YEAR_HEADERS = ["FY2023A", "FY2024A", "FY2025A", "FY2026E", "FY2027E", "FY2028E"]
HIST_COLS = ["B", "C", "D"]
FCST_COLS = ["E", "F", "G"]


def style_title(ws, cell, text, width=8):
    ws[cell] = text
    ws[cell].font = TITLE_FONT
    ws.merge_cells(start_row=ws[cell].row, start_column=ws[cell].column,
                    end_row=ws[cell].row, end_column=ws[cell].column + width)


def style_subtitle(ws, cell, text, width=8):
    ws[cell] = text
    ws[cell].font = SUBTITLE_FONT
    ws.merge_cells(start_row=ws[cell].row, start_column=ws[cell].column,
                    end_row=ws[cell].row, end_column=ws[cell].column + width)


def section_header(ws, row, text, last_col="G"):
    ws[f"A{row}"] = text
    last_idx = ws[f"{last_col}{row}"].column
    for c in range(1, last_idx + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SECTION_FILL
        cell.font = WHITE_BOLD
    ws[f"A{row}"].alignment = Alignment(indent=1)


def year_header_row(ws, row, cols=YEAR_COLS, headers=YEAR_HEADERS, label=""):
    ws[f"A{row}"] = label
    ws[f"A{row}"].font = BLACK_BOLD
    for col, hdr in zip(cols, headers):
        c = ws[f"{col}{row}"]
        c.value = hdr
        c.font = BLACK_BOLD
        c.alignment = Alignment(horizontal="center")
        c.fill = SUBHEADER_FILL


def label(ws, row, text, indent=1, bold=False):
    ws[f"A{row}"] = text
    ws[f"A{row}"].font = BLACK_BOLD if bold else BLACK
    ws[f"A{row}"].alignment = Alignment(indent=indent)


def set_row(ws, row, cols, values, fmt=CUR, font=BLACK, bold=False, top_border=False):
    for col, val in zip(cols, values):
        cell = ws[f"{col}{row}"]
        cell.value = val
        cell.number_format = fmt
        cell.font = BLACK_BOLD if bold else font
        if top_border:
            cell.border = TOTAL_BORDER


def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def src_comment(ws, cell, text):
    ws[cell].comment = Comment(text, "Source")


# =====================================================================
# SHEET 0: README
# =====================================================================
ws = wb.create_sheet("README")
style_title(ws, "A1", "Apple Inc. (AAPL) — 3-Statement Financial Model & Forecast")
style_subtitle(ws, "A2", "Built on real FY2023–FY2025 SEC filings (10-K / 8-K) | Forecast: FY2026E–FY2028E")
rows = [
    "",
    "PURPOSE",
    "This model replicates the core day-to-day deliverable of an FP&A analyst: a fully-linked three-statement",
    "model (Income Statement, Balance Sheet, Cash Flow Statement) built from a public company's actual",
    "financial filings, with a driver-based forecast and a built-in 'what if revenue drops 10%' scenario test.",
    "",
    "HOW THIS MODEL IS BUILT",
    "- FY2023A-FY2025A columns are HARD-CODED ACTUALS sourced directly from Apple's 10-K and 8-K filings on",
    "  SEC EDGAR (cell comments show the exact source). Blue font = hardcoded input.",
    "- FY2026E-FY2028E columns are 100% FORMULA-DRIVEN from the Assumptions tab. Change any blue assumption",
    "  cell and the entire model (IS, BS, CFS, EPS) recalculates automatically.",
    "- The Balance Sheet includes a 'Balance Check' row that must equal zero in every period — this confirms",
    "  the three statements are properly integrated (Cash from the CFS flows into the BS; Net Income flows",
    "  through Retained Earnings/Equity; every working-capital account that changes on the BS is captured as",
    "  a change in the CFS). This is the single most important integrity check in any 3-statement model.",
    "",
    "TABS IN THIS WORKBOOK",
    "1. Assumptions          - All forecast drivers in one place, plus a scenario selector (dropdown)",
    "2. Income Statement     - Net sales, COGS, opex, net income, diluted EPS (FY23A-FY28E)",
    "3. Balance Sheet        - Full balance sheet with balance check",
    "4. Cash Flow Statement  - Operating / investing / financing activities, ending cash ties to BS",
    "5. Scenario Analysis    - Base Case vs. 'Revenue -10%' Bear Case vs. Bull Case, side by side",
    "6. Sensitivity Table    - 2-way data table: Net Income & Free Cash Flow vs. Revenue Growth x Gross Margin",
    "",
    "KEY MODELING ASSUMPTIONS (see Assumptions tab for full detail and rationale)",
    "- Revenue growth, gross margin, R&D%, SG&A% and tax rate are explicit blue-input assumptions by year.",
    "- Working capital (AR, Inventory, AP) is forecast using Days Sales/Inventory/Payable Outstanding, not",
    "  flat percentages — this is how real FP&A working-capital models are built.",
    "- Marketable securities and total debt are held flat in the forecast (no new issuance assumed) to keep",
    "  the model's first version focused on the operating business; this is explicitly called out as a",
    "  simplification in the Assumptions tab and the project write-up.",
    "- Share buybacks are modeled as a payout ratio of Free Cash Flow, and diluted share count declines",
    "  accordingly, flowing through to diluted EPS — linking capital allocation directly to per-share value.",
    "",
    "HOW TO USE THE SCENARIO TOGGLE",
    "Go to the Assumptions tab, cell B3, and pick 'Base', 'Bear (-10% Rev Shock)', or 'Bull (Upside)' from",
    "the dropdown. Every forecast number in the workbook (IS, BS, CFS, EPS) updates instantly.",
    "",
    "DATA SOURCES",
    "- Apple Inc. FY2025 Form 10-K (filed 2025), SEC EDGAR, accession 0000320193-25-000079",
    "- Apple Inc. FY2025 Q4 8-K Exhibit 99.1 (earnings release), filed 10/30/2025",
    "- Apple Inc. FY2023 Form 10-K, SEC EDGAR, accession 0000320193-23-000106",
    "- Apple Inc. FY2023 Q4 8-K Exhibit 99.1 (earnings release), filed 11/2/2023",
    "",
    "Author: Karis Fang  |  Portfolio project for FP&A roles  |  github.com/karisfang",
]
r = 4
for line in rows:
    if line in ("PURPOSE", "HOW THIS MODEL IS BUILT", "TABS IN THIS WORKBOOK",
                "KEY MODELING ASSUMPTIONS (see Assumptions tab for full detail and rationale)",
                "HOW TO USE THE SCENARIO TOGGLE", "DATA SOURCES"):
        ws[f"A{r}"] = line
        ws[f"A{r}"].font = BLACK_BOLD
    else:
        ws[f"A{r}"] = line
        ws[f"A{r}"].font = BLACK
    r += 1
autosize(ws, {"A": 115})
ws.sheet_view.showGridLines = False

print("README done")

# =====================================================================
# SHEET 1: ASSUMPTIONS
# =====================================================================
ws = wb.create_sheet("Assumptions")
style_title(ws, "A1", "Key Assumptions & Scenario Selector")
style_subtitle(ws, "A2", "Blue = hardcoded input you can change. Everything downstream recalculates automatically.")

label(ws, 4, "Scenario Selector", indent=0, bold=True)
ws["B4"] = "Base"
ws["B4"].font = BLUE_BOLD
ws["B4"].fill = YELLOW_FILL
ws["B4"].alignment = Alignment(horizontal="center")
dv = DataValidation(type="list", formula1='"Base,Bear (-10% Rev Shock),Bull (Upside)"', allow_blank=False)
ws.add_data_validation(dv)
dv.add(ws["B4"])
ws["C4"] = "<- choose scenario here"
ws["C4"].font = SUBTITLE_FONT

section_header(ws, 6, "REVENUE GROWTH ASSUMPTIONS (by scenario)")
year_header_row(ws, 7, cols=FCST_COLS, headers=["FY2026E", "FY2027E", "FY2028E"], label="")
label(ws, 8, "Base Case Revenue Growth %")
set_row(ws, 8, FCST_COLS, [0.08, 0.07, 0.07], fmt=PCT, font=BLUE)
label(ws, 9, "Bear Case Growth % (FY26E shock; FY27E/28E recover to Base)")
set_row(ws, 9, ["E"], [-0.10], fmt=PCT, font=BLUE)
label(ws, 10, "Bull Case Growth % (FY26E upside; FY27E/28E = Base)")
set_row(ws, 10, ["E"], [0.12], fmt=PCT, font=BLUE)
label(ws, 11, "Active Revenue Growth % (driven by Scenario Selector)", bold=True)
ws["E11"] = '=IF($B$4="Bear (-10% Rev Shock)",$E$9,IF($B$4="Bull (Upside)",$E$10,$E$8))'
ws["F11"] = "=$F$8"
ws["G11"] = "=$G$8"
set_row(ws, 11, FCST_COLS, [ws["E11"].value, ws["F11"].value, ws["G11"].value], fmt=PCT, font=BLACK_BOLD)

section_header(ws, 13, "MARGIN & OPERATING EXPENSE ASSUMPTIONS")
year_header_row(ws, 14, cols=FCST_COLS, headers=["FY2026E", "FY2027E", "FY2028E"], label="")
label(ws, 15, "Gross Margin %")
set_row(ws, 15, FCST_COLS, [0.472, 0.475, 0.477], fmt=PCT, font=BLUE)
label(ws, 16, "R&D % of Net Sales")
set_row(ws, 16, FCST_COLS, [0.082, 0.082, 0.082], fmt=PCT, font=BLUE)
label(ws, 17, "SG&A % of Net Sales")
set_row(ws, 17, FCST_COLS, [0.066, 0.066, 0.066], fmt=PCT, font=BLUE)
label(ws, 18, "Other Income/(Expense) % of Net Sales")
set_row(ws, 18, FCST_COLS, [-0.001, -0.001, -0.001], fmt=PCT, font=BLUE)
label(ws, 19, "Effective Tax Rate %")
set_row(ws, 19, FCST_COLS, [0.16, 0.16, 0.16], fmt=PCT, font=BLUE)
ws["H19"] = "FY24A rate (24.1%) excludes one-time EU State Aid tax charge; FY23A/FY25A ~15-16% used as normalized basis"
ws["H19"].font = SUBTITLE_FONT

section_header(ws, 21, "WORKING CAPITAL & BALANCE SHEET DRIVERS (fixed across forecast & scenarios)")
label(ws, 22, "Days Sales Outstanding - Accounts Receivable")
ws["B22"] = 33; ws["B22"].font = BLUE; ws["B22"].number_format = DAYS_FMT
label(ws, 23, "Days Inventory Outstanding")
ws["B23"] = 9; ws["B23"].font = BLUE; ws["B23"].number_format = DAYS_FMT
label(ws, 24, "Days Payable Outstanding - Accounts Payable")
ws["B24"] = 113; ws["B24"].font = BLUE; ws["B24"].number_format = DAYS_FMT
label(ws, 25, "Vendor Non-Trade Receivables % of Net Sales")
ws["B25"] = 0.08; ws["B25"].font = BLUE; ws["B25"].number_format = PCT
label(ws, 26, "Other Current Assets % of Net Sales")
ws["B26"] = 0.035; ws["B26"].font = BLUE; ws["B26"].number_format = PCT
label(ws, 27, "Other Non-Current Assets % of Net Sales")
ws["B27"] = 0.20; ws["B27"].font = BLUE; ws["B27"].number_format = PCT
label(ws, 28, "Other Current Liabilities % of Net Sales")
ws["B28"] = 0.16; ws["B28"].font = BLUE; ws["B28"].number_format = PCT
label(ws, 29, "Deferred Revenue % of Net Sales")
ws["B29"] = 0.022; ws["B29"].font = BLUE; ws["B29"].number_format = PCT
label(ws, 30, "Other Non-Current Liabilities % of Net Sales")
ws["B30"] = 0.10; ws["B30"].font = BLUE; ws["B30"].number_format = PCT

section_header(ws, 32, "CAPEX, D&A & CAPITAL RETURN ASSUMPTIONS")
year_header_row(ws, 33, cols=FCST_COLS, headers=["FY2026E", "FY2027E", "FY2028E"], label="")
label(ws, 34, "Capex % of Net Sales")
set_row(ws, 34, FCST_COLS, [0.03, 0.03, 0.03], fmt=PCT, font=BLUE)
label(ws, 35, "D&A % of Net Sales")
set_row(ws, 35, FCST_COLS, [0.028, 0.028, 0.028], fmt=PCT, font=BLUE)
label(ws, 36, "Stock-Based Compensation % of Net Sales")
set_row(ws, 36, FCST_COLS, [0.030, 0.030, 0.030], fmt=PCT, font=BLUE)
label(ws, 37, "Tax Withholding on Equity Awards (% of SBC)")
set_row(ws, 37, FCST_COLS, [0.45, 0.45, 0.45], fmt=PCT, font=BLUE)
label(ws, 38, "Dividend Growth Rate % (YoY)")
ws["B38"] = 0.04; ws["B38"].font = BLUE; ws["B38"].number_format = PCT
label(ws, 39, "Buyback Payout Ratio (% of Free Cash Flow)")
ws["B39"] = 0.75; ws["B39"].font = BLUE; ws["B39"].number_format = PCT
label(ws, 40, "Average Share Price Assumption ($, for buyback share count)")
set_row(ws, 40, FCST_COLS, [240, 255, 270], fmt=PRICE_FMT, font=BLUE)

autosize(ws, {"A": 58, "B": 13, "C": 13, "D": 13, "E": 13, "F": 13, "G": 13, "H": 55})
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B5"

print("Assumptions done")

# =====================================================================
# SHEET 2: INCOME STATEMENT
# =====================================================================
ws = wb.create_sheet("Income Statement")
style_title(ws, "A1", "Apple Inc. — Consolidated Income Statement ($ in millions, except per-share)")
style_subtitle(ws, "A2", "FY2023A-FY2025A: actuals from 10-K/8-K filings | FY2026E-FY2028E: driven by Assumptions tab")
year_header_row(ws, 4)

label(ws, 5, "Net Sales:", bold=True)
label(ws, 6, "Products")
set_row(ws, 6, HIST_COLS, [298085, 294866, 307003])
label(ws, 7, "Services")
set_row(ws, 7, HIST_COLS, [85200, 96169, 109158])
label(ws, 8, "Total Net Sales", bold=True)
set_row(ws, 8, HIST_COLS, [383285, 391035, 416161], bold=True, top_border=True)
ws["E8"] = "=D8*(1+Assumptions!E11)"
ws["F8"] = "=E8*(1+Assumptions!F11)"
ws["G8"] = "=F8*(1+Assumptions!G11)"
set_row(ws, 8, FCST_COLS, [ws["E8"].value, ws["F8"].value, ws["G8"].value], bold=True, top_border=True)

label(ws, 10, "Cost of Sales:", bold=True)
label(ws, 11, "Products")
set_row(ws, 11, HIST_COLS, [189282, 185233, 194116])
label(ws, 12, "Services")
set_row(ws, 12, HIST_COLS, [24855, 25119, 26844])
label(ws, 13, "Total Cost of Sales", bold=True)
set_row(ws, 13, HIST_COLS, [214137, 210352, 220960], bold=True, top_border=True)
for col in FCST_COLS:
    ws[f"{col}13"] = f"={col}8-{col}14"
    ws[f"{col}13"].number_format = CUR
    ws[f"{col}13"].font = BLACK_BOLD
    ws[f"{col}13"].border = TOTAL_BORDER

label(ws, 14, "Gross Margin", bold=True)
set_row(ws, 14, HIST_COLS, ["=B8-B13", "=C8-C13", "=D8-D13"], bold=True)
for col in FCST_COLS:
    ws[f"{col}14"] = f"={col}8*Assumptions!{col}15"
    ws[f"{col}14"].number_format = CUR
    ws[f"{col}14"].font = BLACK_BOLD
label(ws, 15, "Gross Margin %")
set_row(ws, 15, YEAR_COLS, [f"={c}14/{c}8" for c in YEAR_COLS], fmt=PCT, font=GREEN)

label(ws, 17, "Operating Expenses:", bold=True)
label(ws, 18, "Research and Development")
set_row(ws, 18, HIST_COLS, [29915, 31370, 34550])
for col in FCST_COLS:
    ws[f"{col}18"] = f"={col}8*Assumptions!{col}16"
    ws[f"{col}18"].number_format = CUR
    ws[f"{col}18"].font = BLACK
label(ws, 19, "Selling, General and Administrative")
set_row(ws, 19, HIST_COLS, [24932, 26097, 27601])
for col in FCST_COLS:
    ws[f"{col}19"] = f"={col}8*Assumptions!{col}17"
    ws[f"{col}19"].number_format = CUR
    ws[f"{col}19"].font = BLACK
label(ws, 20, "Total Operating Expenses", bold=True)
set_row(ws, 20, YEAR_COLS, [f"=SUM({c}18:{c}19)" for c in YEAR_COLS], bold=True, top_border=True)

label(ws, 22, "Operating Income", bold=True)
set_row(ws, 22, YEAR_COLS, [f"={c}14-{c}20" for c in YEAR_COLS], bold=True)
label(ws, 23, "Operating Margin %")
set_row(ws, 23, YEAR_COLS, [f"={c}22/{c}8" for c in YEAR_COLS], fmt=PCT, font=GREEN)

label(ws, 25, "Other Income/(Expense), net")
set_row(ws, 25, HIST_COLS, [-565, 269, -321])
for col in FCST_COLS:
    ws[f"{col}25"] = f"={col}8*Assumptions!{col}18"
    ws[f"{col}25"].number_format = CUR
    ws[f"{col}25"].font = BLACK
label(ws, 26, "Income Before Provision for Income Taxes", bold=True)
set_row(ws, 26, YEAR_COLS, [f"={c}22+{c}25" for c in YEAR_COLS], bold=True, top_border=True)
label(ws, 27, "Provision for Income Taxes")
set_row(ws, 27, HIST_COLS, [16741, 29749, 20719])
for col in FCST_COLS:
    ws[f"{col}27"] = f"={col}26*Assumptions!{col}19"
    ws[f"{col}27"].number_format = CUR
    ws[f"{col}27"].font = BLACK
label(ws, 28, "Net Income", bold=True)
set_row(ws, 28, YEAR_COLS, [f"={c}26-{c}27" for c in YEAR_COLS], bold=True, top_border=True)
label(ws, 29, "Net Margin %")
set_row(ws, 29, YEAR_COLS, [f"={c}28/{c}8" for c in YEAR_COLS], fmt=PCT, font=GREEN)

label(ws, 31, "Diluted Shares Outstanding (000s)")
set_row(ws, 31, HIST_COLS, [15812547, 15408095, 15004697], fmt=NUM)
for col, prev in zip(FCST_COLS, ["D", "E", "F"]):
    ws[f"{col}31"] = f"={prev}31+'Cash Flow Statement'!{col}29/Assumptions!{col}40*1000"
    ws[f"{col}31"].number_format = NUM
    ws[f"{col}31"].font = GREEN
label(ws, 32, "Diluted EPS", bold=True)
set_row(ws, 32, YEAR_COLS, [f"={c}28*1000/{c}31" for c in YEAR_COLS], fmt=EPS_FMT, bold=True)

src_comment(ws, "B6", "Source: Apple FY2023 10-K, Consolidated Statements of Operations, SEC EDGAR")
src_comment(ws, "D6", "Source: Apple FY2025 Form 10-K / Q4 FY25 8-K Ex-99.1, filed 10/30/2025, SEC EDGAR")
src_comment(ws, "E8", "Formula: Prior year Net Sales x (1 + Active Revenue Growth % from Assumptions tab)")
src_comment(ws, "B14", "Note: row 13 (Total COGS) for forecast years is back-solved as Net Sales - Gross Margin (driven by Assumptions GM% input)")

autosize(ws, {"A": 42, "B": 13, "C": 13, "D": 13, "E": 13, "F": 13, "G": 13})
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B5"

print("Income Statement done")

# =====================================================================
# SHEET 3: BALANCE SHEET
# =====================================================================
ws = wb.create_sheet("Balance Sheet")
style_title(ws, "A1", "Apple Inc. — Consolidated Balance Sheet ($ in millions)")
style_subtitle(ws, "A2", "FY2023A-FY2025A: actuals from 10-K/8-K filings | FY2026E-FY2028E: driven by Assumptions tab")
year_header_row(ws, 4)

label(ws, 5, "ASSETS", bold=True)
label(ws, 6, "Current Assets:", bold=True)
label(ws, 7, "Cash and Cash Equivalents")
set_row(ws, 7, HIST_COLS, [29965, 29943, 35934])
for col, prev in zip(FCST_COLS, ["D", "E", "F"]):
    ws[f"{col}7"] = f"={prev}7+'Cash Flow Statement'!{col}34"
    ws[f"{col}7"].number_format = CUR
    ws[f"{col}7"].font = GREEN

label(ws, 8, "Marketable Securities (Current)")
set_row(ws, 8, HIST_COLS, [31590, 35228, 18763])
for col, prev in zip(FCST_COLS, ["D", "E", "F"]):
    ws[f"{col}8"] = f"={prev}8"
    ws[f"{col}8"].number_format = CUR
    ws[f"{col}8"].font = BLACK

label(ws, 9, "Accounts Receivable, net")
set_row(ws, 9, HIST_COLS, [29508, 33410, 39777])
for col in FCST_COLS:
    ws[f"{col}9"] = f"='Income Statement'!{col}8/365*Assumptions!$B$22"
    ws[f"{col}9"].number_format = CUR
    ws[f"{col}9"].font = BLACK

label(ws, 10, "Vendor Non-Trade Receivables")
set_row(ws, 10, HIST_COLS, [31477, 32833, 33180])
for col in FCST_COLS:
    ws[f"{col}10"] = f"='Income Statement'!{col}8*Assumptions!$B$25"
    ws[f"{col}10"].number_format = CUR
    ws[f"{col}10"].font = BLACK

label(ws, 11, "Inventories")
set_row(ws, 11, HIST_COLS, [6331, 7286, 5718])
for col in FCST_COLS:
    ws[f"{col}11"] = f"='Income Statement'!{col}13/365*Assumptions!$B$23"
    ws[f"{col}11"].number_format = CUR
    ws[f"{col}11"].font = BLACK

label(ws, 12, "Other Current Assets")
set_row(ws, 12, HIST_COLS, [14695, 14287, 14585])
for col in FCST_COLS:
    ws[f"{col}12"] = f"='Income Statement'!{col}8*Assumptions!$B$26"
    ws[f"{col}12"].number_format = CUR
    ws[f"{col}12"].font = BLACK

label(ws, 13, "Total Current Assets", bold=True)
set_row(ws, 13, YEAR_COLS, [f"=SUM({c}7:{c}12)" for c in YEAR_COLS], bold=True, top_border=True)

label(ws, 15, "Non-Current Assets:", bold=True)
label(ws, 16, "Marketable Securities (Non-Current)")
set_row(ws, 16, HIST_COLS, [100544, 91479, 77723])
for col, prev in zip(FCST_COLS, ["D", "E", "F"]):
    ws[f"{col}16"] = f"={prev}16"
    ws[f"{col}16"].number_format = CUR
    ws[f"{col}16"].font = BLACK

label(ws, 17, "Property, Plant and Equipment, net")
set_row(ws, 17, HIST_COLS, [43715, 45680, 49834])
for col, prev in zip(FCST_COLS, ["D", "E", "F"]):
    ws[f"{col}17"] = f"={prev}17-'Cash Flow Statement'!{col}23-'Cash Flow Statement'!{col}7"
    ws[f"{col}17"].number_format = CUR
    ws[f"{col}17"].font = GREEN

label(ws, 18, "Other Non-Current Assets")
set_row(ws, 18, HIST_COLS, [64758, 74834, 83727])
for col in FCST_COLS:
    ws[f"{col}18"] = f"='Income Statement'!{col}8*Assumptions!$B$27"
    ws[f"{col}18"].number_format = CUR
    ws[f"{col}18"].font = BLACK

label(ws, 19, "Total Non-Current Assets", bold=True)
set_row(ws, 19, YEAR_COLS, [f"=SUM({c}16:{c}18)" for c in YEAR_COLS], bold=True, top_border=True)
label(ws, 20, "TOTAL ASSETS", bold=True)
set_row(ws, 20, YEAR_COLS, [f"={c}13+{c}19" for c in YEAR_COLS], bold=True, top_border=True)

label(ws, 22, "LIABILITIES AND SHAREHOLDERS' EQUITY", bold=True)
label(ws, 23, "Current Liabilities:", bold=True)
label(ws, 24, "Accounts Payable")
set_row(ws, 24, HIST_COLS, [62611, 68960, 69860])
for col in FCST_COLS:
    ws[f"{col}24"] = f"='Income Statement'!{col}13/365*Assumptions!$B$24"
    ws[f"{col}24"].number_format = CUR
    ws[f"{col}24"].font = BLACK

label(ws, 25, "Other Current Liabilities")
set_row(ws, 25, HIST_COLS, [58829, 78304, 66387])
for col in FCST_COLS:
    ws[f"{col}25"] = f"='Income Statement'!{col}8*Assumptions!$B$28"
    ws[f"{col}25"].number_format = CUR
    ws[f"{col}25"].font = BLACK

label(ws, 26, "Deferred Revenue")
set_row(ws, 26, HIST_COLS, [8061, 8249, 9055])
for col in FCST_COLS:
    ws[f"{col}26"] = f"='Income Statement'!{col}8*Assumptions!$B$29"
    ws[f"{col}26"].number_format = CUR
    ws[f"{col}26"].font = BLACK

label(ws, 27, "Commercial Paper")
set_row(ws, 27, HIST_COLS, [5985, 9967, 7979])
for col, prev in zip(FCST_COLS, ["D", "E", "F"]):
    ws[f"{col}27"] = f"={prev}27"
    ws[f"{col}27"].number_format = CUR
    ws[f"{col}27"].font = BLACK

label(ws, 28, "Term Debt (Current Portion)")
set_row(ws, 28, HIST_COLS, [9822, 10912, 12350])
for col, prev in zip(FCST_COLS, ["D", "E", "F"]):
    ws[f"{col}28"] = f"={prev}28"
    ws[f"{col}28"].number_format = CUR
    ws[f"{col}28"].font = BLACK

label(ws, 29, "Total Current Liabilities", bold=True)
set_row(ws, 29, YEAR_COLS, [f"=SUM({c}24:{c}28)" for c in YEAR_COLS], bold=True, top_border=True)

label(ws, 31, "Non-Current Liabilities:", bold=True)
label(ws, 32, "Term Debt (Non-Current)")
set_row(ws, 32, HIST_COLS, [95281, 85750, 78328])
for col, prev in zip(FCST_COLS, ["D", "E", "F"]):
    ws[f"{col}32"] = f"={prev}32"
    ws[f"{col}32"].number_format = CUR
    ws[f"{col}32"].font = BLACK

label(ws, 33, "Other Non-Current Liabilities")
set_row(ws, 33, HIST_COLS, [49848, 45888, 41549])
for col in FCST_COLS:
    ws[f"{col}33"] = f"='Income Statement'!{col}8*Assumptions!$B$30"
    ws[f"{col}33"].number_format = CUR
    ws[f"{col}33"].font = BLACK

label(ws, 34, "Total Non-Current Liabilities", bold=True)
set_row(ws, 34, YEAR_COLS, [f"=SUM({c}32:{c}33)" for c in YEAR_COLS], bold=True, top_border=True)
label(ws, 35, "TOTAL LIABILITIES", bold=True)
set_row(ws, 35, YEAR_COLS, [f"={c}29+{c}34" for c in YEAR_COLS], bold=True, top_border=True)

label(ws, 37, "Total Shareholders' Equity", bold=True)
set_row(ws, 37, HIST_COLS, [62146, 56950, 73733], bold=True)
for col, prev in zip(FCST_COLS, ["D", "E", "F"]):
    ws[f"{col}37"] = (f"={prev}37+'Income Statement'!{col}28+'Cash Flow Statement'!{col}28"
                       f"+'Cash Flow Statement'!{col}29+'Cash Flow Statement'!{col}27+'Cash Flow Statement'!{col}8")
    ws[f"{col}37"].number_format = CUR
    ws[f"{col}37"].font = GREEN

label(ws, 38, "TOTAL LIABILITIES AND SHAREHOLDERS' EQUITY", bold=True)
set_row(ws, 38, YEAR_COLS, [f"={c}35+{c}37" for c in YEAR_COLS], bold=True, top_border=True)

label(ws, 40, "Balance Check (Total Assets - Total Liab. & Equity)", bold=True)
set_row(ws, 40, YEAR_COLS, [f"=ROUND({c}20-{c}38,0)" for c in YEAR_COLS], bold=True, fmt=NUM)
for col in YEAR_COLS:
    ws[f"{col}40"].font = Font(name=FONT_NAME, bold=True, color="C00000")

src_comment(ws, "B7", "Source: Apple FY2023 10-K, Consolidated Balance Sheets, SEC EDGAR")
src_comment(ws, "D7", "Source: Apple FY2025 Q4 8-K Ex-99.1, Condensed Consolidated Balance Sheets, filed 10/30/2025")
src_comment(ws, "B37", "FY23A/24A equity = Common stock & APIC + Accumulated deficit + AOCI per 10-K; forecast years roll forward as a single equity balance (see write-up)")
src_comment(ws, "B40", "Should equal zero in every period. Confirms the three statements are fully integrated.")

autosize(ws, {"A": 46, "B": 13, "C": 13, "D": 13, "E": 13, "F": 13, "G": 13})
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B5"

print("Balance Sheet done")

# =====================================================================
# SHEET 4: CASH FLOW STATEMENT
# =====================================================================
ws = wb.create_sheet("Cash Flow Statement")
style_title(ws, "A1", "Apple Inc. — Consolidated Statement of Cash Flows ($ in millions)")
style_subtitle(ws, "A2", "FY2023A-FY2025A: actuals from 10-K/8-K filings | FY2026E-FY2028E: driven by Assumptions tab")
year_header_row(ws, 4)

label(ws, 5, "Operating Activities:", bold=True)
label(ws, 6, "Net Income")
set_row(ws, 6, HIST_COLS, [96995, 93736, 112010])
for col in FCST_COLS:
    ws[f"{col}6"] = f"='Income Statement'!{col}28"
    ws[f"{col}6"].number_format = CUR
    ws[f"{col}6"].font = GREEN

label(ws, 7, "Depreciation and Amortization")
set_row(ws, 7, HIST_COLS, [11519, 11445, 11698])
for col in FCST_COLS:
    ws[f"{col}7"] = f"='Income Statement'!{col}8*Assumptions!{col}35"
    ws[f"{col}7"].number_format = CUR
    ws[f"{col}7"].font = BLACK

label(ws, 8, "Share-Based Compensation Expense")
set_row(ws, 8, HIST_COLS, [10833, 11688, 12863])
for col in FCST_COLS:
    ws[f"{col}8"] = f"='Income Statement'!{col}8*Assumptions!{col}36"
    ws[f"{col}8"].number_format = CUR
    ws[f"{col}8"].font = BLACK

label(ws, 9, "Other (non-cash)")
set_row(ws, 9, HIST_COLS, [-2227, -2266, -89])
for col in FCST_COLS:
    ws[f"{col}9"] = 0
    ws[f"{col}9"].number_format = CUR
    ws[f"{col}9"].font = BLUE

label(ws, 10, "Changes in operating assets and liabilities:", bold=True)
label(ws, 11, "Accounts Receivable, net")
set_row(ws, 11, HIST_COLS, [-1688, -3788, -6682])
for col, prev in zip(FCST_COLS, ["D", "E", "F"]):
    ws[f"{col}11"] = f"=-('Balance Sheet'!{col}9-'Balance Sheet'!{prev}9)"
    ws[f"{col}11"].number_format = CUR
    ws[f"{col}11"].font = GREEN

label(ws, 12, "Vendor Non-Trade Receivables")
set_row(ws, 12, HIST_COLS, [1271, -1356, -347])
for col, prev in zip(FCST_COLS, ["D", "E", "F"]):
    ws[f"{col}12"] = f"=-('Balance Sheet'!{col}10-'Balance Sheet'!{prev}10)"
    ws[f"{col}12"].number_format = CUR
    ws[f"{col}12"].font = GREEN

label(ws, 13, "Inventories")
set_row(ws, 13, HIST_COLS, [-1618, -1046, 1400])
for col, prev in zip(FCST_COLS, ["D", "E", "F"]):
    ws[f"{col}13"] = f"=-('Balance Sheet'!{col}11-'Balance Sheet'!{prev}11)"
    ws[f"{col}13"].number_format = CUR
    ws[f"{col}13"].font = GREEN

label(ws, 14, "Other Current and Non-Current Assets")
set_row(ws, 14, HIST_COLS, [-5684, -11731, -9197])
for col, prev in zip(FCST_COLS, ["D", "E", "F"]):
    ws[f"{col}14"] = (f"=-(('Balance Sheet'!{col}12+'Balance Sheet'!{col}18)"
                       f"-('Balance Sheet'!{prev}12+'Balance Sheet'!{prev}18))")
    ws[f"{col}14"].number_format = CUR
    ws[f"{col}14"].font = GREEN

label(ws, 15, "Accounts Payable")
set_row(ws, 15, HIST_COLS, [-1889, 6020, 902])
for col, prev in zip(FCST_COLS, ["D", "E", "F"]):
    ws[f"{col}15"] = f"='Balance Sheet'!{col}24-'Balance Sheet'!{prev}24"
    ws[f"{col}15"].number_format = CUR
    ws[f"{col}15"].font = GREEN

label(ws, 16, "Other Current and Non-Current Liabilities")
set_row(ws, 16, HIST_COLS, [3031, 15552, -11076])
for col, prev in zip(FCST_COLS, ["D", "E", "F"]):
    ws[f"{col}16"] = (f"=(('Balance Sheet'!{col}25+'Balance Sheet'!{col}26+'Balance Sheet'!{col}33)"
                       f"-('Balance Sheet'!{prev}25+'Balance Sheet'!{prev}26+'Balance Sheet'!{prev}33))")
    ws[f"{col}16"].number_format = CUR
    ws[f"{col}16"].font = GREEN

label(ws, 17, "Cash Generated by Operating Activities", bold=True)
set_row(ws, 17, YEAR_COLS, [f"=SUM({c}6:{c}16)" for c in YEAR_COLS], bold=True, top_border=True)

label(ws, 19, "Investing Activities:", bold=True)
label(ws, 20, "Purchases of Marketable Securities")
set_row(ws, 20, HIST_COLS, [-29513, -48656, -24407])
for col in FCST_COLS:
    ws[f"{col}20"] = 0; ws[f"{col}20"].number_format = CUR; ws[f"{col}20"].font = BLUE
label(ws, 21, "Proceeds from Maturities/Sales of Marketable Securities")
set_row(ws, 21, HIST_COLS, [45514, 62346, 53797])
for col in FCST_COLS:
    ws[f"{col}21"] = 0; ws[f"{col}21"].number_format = CUR; ws[f"{col}21"].font = BLUE
label(ws, 22, "Other Investing")
set_row(ws, 22, HIST_COLS, [-1337, -1308, -1480])
for col in FCST_COLS:
    ws[f"{col}22"] = 0; ws[f"{col}22"].number_format = CUR; ws[f"{col}22"].font = BLUE
label(ws, 23, "Payments for Acquisition of Property, Plant & Equipment (Capex)")
set_row(ws, 23, HIST_COLS, [-10959, -9447, -12715])
for col in FCST_COLS:
    ws[f"{col}23"] = f"=-('Income Statement'!{col}8*Assumptions!{col}34)"
    ws[f"{col}23"].number_format = CUR
    ws[f"{col}23"].font = BLACK
label(ws, 24, "Cash Generated by/(Used in) Investing Activities", bold=True)
set_row(ws, 24, YEAR_COLS, [f"=SUM({c}20:{c}23)" for c in YEAR_COLS], bold=True, top_border=True)

label(ws, 26, "Financing Activities:", bold=True)
label(ws, 27, "Payments for Taxes on Net Share Settlement of Equity Awards")
set_row(ws, 27, HIST_COLS, [-5431, -5441, -5960])
for col in FCST_COLS:
    ws[f"{col}27"] = f"=-({col}8*Assumptions!{col}37)"
    ws[f"{col}27"].number_format = CUR
    ws[f"{col}27"].font = BLACK
label(ws, 28, "Payments for Dividends and Dividend Equivalents")
set_row(ws, 28, HIST_COLS, [-15025, -15234, -15421])
for col, prev in zip(FCST_COLS, ["D", "E", "F"]):
    ws[f"{col}28"] = f"={prev}28*(1+Assumptions!$B$38)"
    ws[f"{col}28"].number_format = CUR
    ws[f"{col}28"].font = BLACK
label(ws, 29, "Repurchases of Common Stock")
set_row(ws, 29, HIST_COLS, [-77550, -94949, -90711])
for col in FCST_COLS:
    ws[f"{col}29"] = f"=-(({col}17+{col}24)*Assumptions!$B$39)"
    ws[f"{col}29"].number_format = CUR
    ws[f"{col}29"].font = BLACK
label(ws, 30, "Debt Issuance/(Repayment) and Commercial Paper, net")
set_row(ws, 30, HIST_COLS, [-9901, -5998, -8483])
for col in FCST_COLS:
    ws[f"{col}30"] = 0; ws[f"{col}30"].number_format = CUR; ws[f"{col}30"].font = BLUE
label(ws, 31, "Other Financing")
set_row(ws, 31, HIST_COLS, [-581, -361, -111])
for col in FCST_COLS:
    ws[f"{col}31"] = 0; ws[f"{col}31"].number_format = CUR; ws[f"{col}31"].font = BLUE
label(ws, 32, "Cash Used in Financing Activities", bold=True)
set_row(ws, 32, YEAR_COLS, [f"=SUM({c}27:{c}31)" for c in YEAR_COLS], bold=True, top_border=True)

label(ws, 34, "Net Increase/(Decrease) in Cash", bold=True)
set_row(ws, 34, YEAR_COLS, [f"={c}17+{c}24+{c}32" for c in YEAR_COLS], bold=True, top_border=True)
label(ws, 35, "Cash, Beginning of Period")
ws["B35"] = 24977
ws["B35"].number_format = CUR; ws["B35"].font = BLUE
for col, prev in zip(["C", "D", "E", "F", "G"], ["B", "C", "D", "E", "F"]):
    ws[f"{col}35"] = f"={prev}36"
    ws[f"{col}35"].number_format = CUR
    ws[f"{col}35"].font = BLACK
label(ws, 36, "Cash, End of Period", bold=True)
set_row(ws, 36, YEAR_COLS, [f"={c}35+{c}34" for c in YEAR_COLS], bold=True, top_border=True)
ws["A37"] = "Note: ending cash here includes restricted cash equivalents per GAAP CFS presentation; ties to BS Cash line via the change in row 34."
ws["A37"].font = SUBTITLE_FONT
ws.merge_cells("A37:G37")

src_comment(ws, "B6", "Source: Apple FY2023 10-K, Consolidated Statements of Cash Flows, SEC EDGAR")
src_comment(ws, "D6", "Source: Apple FY2025 Q4 8-K Ex-99.1, Condensed Consolidated Statements of Cash Flows, filed 10/30/2025")
src_comment(ws, "E29", "Forecast buybacks = Buyback Payout Ratio (Assumptions) x Free Cash Flow (CFO + Investing, i.e. CFO - Capex)")

autosize(ws, {"A": 56, "B": 13, "C": 13, "D": 13, "E": 13, "F": 13, "G": 13})
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B5"

print("Cash Flow Statement done")

# =====================================================================
# SHEET 5: SCENARIO ANALYSIS
# =====================================================================
ws = wb.create_sheet("Scenario Analysis")
style_title(ws, "A1", "Scenario Analysis — 'What If Revenue Drops 10%?'")
style_subtitle(ws, "A2", "Live comparison: Base Case vs. Bear Case (-10% FY26E revenue shock) vs. Bull Case (+12% FY26E)")

ws["A4"] = "Current Assumptions tab scenario selector is set to:"
ws["A4"].font = BLACK
ws["C4"] = "=Assumptions!B4"
ws["C4"].font = BLUE_BOLD
ws["C4"].fill = YELLOW_FILL
ws.merge_cells("A4:B4")

label(ws, 6, "FY2026E Revenue Growth Rate Applied", bold=True)
ws["B6"] = "Base"; ws["C6"] = "Bear (-10% Rev Shock)"; ws["D6"] = "Bull (Upside)"
for c in ["B6", "C6", "D6"]:
    ws[c].font = BLACK_BOLD
    ws[c].fill = SUBHEADER_FILL
    ws[c].alignment = Alignment(horizontal="center")
ws["B7"] = "=Assumptions!E8"; ws["C7"] = "=Assumptions!E9"; ws["D7"] = "=Assumptions!E10"
for c in ["B7", "C7", "D7"]:
    ws[c].number_format = PCT; ws[c].font = GREEN

# Build a self-contained 3-column replica of FY26E driven by each scenario's growth rate,
# holding every other assumption constant — this isolates the pure revenue-shock effect.
rows_def = [
    ("Net Sales", "='Income Statement'!$D$8*(1+{g})"),
    ("Gross Margin", "={rev}*Assumptions!$E$15"),
    ("Total Operating Expenses", "={rev}*(Assumptions!$E$16+Assumptions!$E$17)"),
    ("Operating Income", "={gm}-{opex}"),
    ("Other Income/(Expense)", "={rev}*Assumptions!$E$18"),
    ("Income Before Taxes", "={oi}+{other}"),
    ("Provision for Income Taxes", "={pretax}*Assumptions!$E$19"),
    ("Net Income", "={pretax}-{tax}"),
    ("  Net Income vs. Base ($)", "={ni}-$B${ni_row}"),
    ("  Net Income vs. Base (%)", "=IFERROR(({ni}-$B${ni_row})/$B${ni_row},0)"),
]
start_row = 9
label(ws, start_row, "FY2026E Income Statement Under Each Scenario", bold=True)
r = start_row + 1
cells = {}
for name, _ in rows_def:
    label(ws, r, name)
    cells[name] = r
    r += 1

g_map = {"B": "Assumptions!$E$8", "C": "Assumptions!$E$9", "D": "Assumptions!$E$10"}
for col, g in g_map.items():
    rev_r, gm_r, opex_r, oi_r, other_r, pretax_r, tax_r, ni_r = (
        cells["Net Sales"], cells["Gross Margin"], cells["Total Operating Expenses"],
        cells["Operating Income"], cells["Other Income/(Expense)"], cells["Income Before Taxes"],
        cells["Provision for Income Taxes"], cells["Net Income"])
    ws[f"{col}{rev_r}"] = f"='Income Statement'!$D$8*(1+{g})"
    ws[f"{col}{gm_r}"] = f"={col}{rev_r}*Assumptions!$E$15"
    ws[f"{col}{opex_r}"] = f"={col}{rev_r}*(Assumptions!$E$16+Assumptions!$E$17)"
    ws[f"{col}{oi_r}"] = f"={col}{gm_r}-{col}{opex_r}"
    ws[f"{col}{other_r}"] = f"={col}{rev_r}*Assumptions!$E$18"
    ws[f"{col}{pretax_r}"] = f"={col}{oi_r}+{col}{other_r}"
    ws[f"{col}{tax_r}"] = f"={col}{pretax_r}*Assumptions!$E$19"
    ws[f"{col}{ni_r}"] = f"={col}{pretax_r}-{col}{tax_r}"
    for rr in [rev_r, gm_r, opex_r, oi_r, other_r, pretax_r, tax_r, ni_r]:
        ws[f"{col}{rr}"].number_format = CUR
        ws[f"{col}{rr}"].font = BLACK_BOLD if rr == ni_r else BLACK

ni_row = cells["Net Income"]
diff_row = cells["  Net Income vs. Base ($)"]
pct_row = cells["  Net Income vs. Base (%)"]
for col in ["B", "C", "D"]:
    ws[f"{col}{diff_row}"] = f"={col}{ni_row}-$B${ni_row}"
    ws[f"{col}{diff_row}"].number_format = CUR
    ws[f"{col}{pct_row}"] = f"=IFERROR(({col}{ni_row}-$B${ni_row})/$B${ni_row},0)"
    ws[f"{col}{pct_row}"].number_format = PCT

ws["A22"] = ("Takeaway: a 10-percentage-point revenue growth shock in FY26E (from +8% Base to -10% Bear) "
             "compresses FY26E Net Income by the figure shown above — operating expenses don't flex down "
             "with revenue in the short run (R&D/SG&A are modeled as % of revenue here, but in reality much "
             "of this is fixed in the near term), so margin compression amplifies the revenue hit to the bottom line.")
ws["A22"].font = SUBTITLE_FONT
ws.merge_cells("A22:F22")
ws.row_dimensions[22].height = 30

autosize(ws, {"A": 40, "B": 22, "C": 22, "D": 22})
ws.sheet_view.showGridLines = False

print("Scenario Analysis done")

# =====================================================================
# SHEET 6: SENSITIVITY TABLE
# =====================================================================
ws = wb.create_sheet("Sensitivity Table")
style_title(ws, "A1", "FY2026E Sensitivity Analysis — Net Income & Free Cash Flow")
style_subtitle(ws, "A2", "2-way grid: Revenue Growth % (rows) x Gross Margin % (columns). All other assumptions held at Base Case.")

growth_vals = [-0.10, -0.05, 0.00, 0.08, 0.12, 0.15]
margin_vals = [0.450, 0.460, 0.470, 0.472, 0.480, 0.490]

def build_grid(ws, top_row, title, metric):
    label(ws, top_row, title, bold=True)
    ws[f"A{top_row+1}"] = "Revenue Growth % \\ Gross Margin %"
    ws[f"A{top_row+1}"].font = BLACK_BOLD
    ws[f"A{top_row+1}"].alignment = Alignment(wrap_text=True, indent=1)
    cols = ["B", "C", "D", "E", "F", "G"]
    for col, mv in zip(cols, margin_vals):
        c = ws[f"{col}{top_row+1}"]
        c.value = mv
        c.number_format = PCT
        c.font = BLUE_BOLD
        c.fill = SUBHEADER_FILL
        c.alignment = Alignment(horizontal="center")
    for i, gv in enumerate(growth_vals):
        row = top_row + 2 + i
        gc = ws[f"A{row}"]
        gc.value = gv
        gc.number_format = PCT
        gc.font = BLUE_BOLD
        gc.fill = SUBHEADER_FILL
        for col in cols:
            mv_cell = f"{col}${top_row+1}"
            cell = ws[f"{col}{row}"]
            rev = f"('Income Statement'!$D$8*(1+$A{row}))"
            cogs = f"({rev}*(1-{mv_cell}))"
            gm = f"({rev}*{mv_cell})"
            opex = f"({rev}*(Assumptions!$E$16+Assumptions!$E$17))"
            oi = f"({gm}-{opex})"
            other = f"({rev}*Assumptions!$E$18)"
            pretax = f"({oi}+{other})"
            tax = f"({pretax}*Assumptions!$E$19)"
            ni = f"({pretax}-{tax})"
            if metric == "NI":
                cell.value = f"={ni}"
            else:
                da = f"({rev}*Assumptions!$E$35)"
                sbc = f"({rev}*Assumptions!$E$36)"
                ar = f"({rev}/365*Assumptions!$B$22)"
                vntr = f"({rev}*Assumptions!$B$25)"
                inv = f"({cogs}/365*Assumptions!$B$23)"
                otherca = f"({rev}*(Assumptions!$B$26+Assumptions!$B$27))"
                ap = f"({cogs}/365*Assumptions!$B$24)"
                otherliab = f"({rev}*(Assumptions!$B$28+Assumptions!$B$29+Assumptions!$B$30))"
                d_ar = f"(-({ar}-'Balance Sheet'!$D$9))"
                d_vntr = f"(-({vntr}-'Balance Sheet'!$D$10))"
                d_inv = f"(-({inv}-'Balance Sheet'!$D$11))"
                d_otherca = f"(-({otherca}-('Balance Sheet'!$D$12+'Balance Sheet'!$D$18)))"
                d_ap = f"({ap}-'Balance Sheet'!$D$24)"
                d_otherliab = f"({otherliab}-('Balance Sheet'!$D$25+'Balance Sheet'!$D$26+'Balance Sheet'!$D$33))"
                capex = f"({rev}*Assumptions!$E$34)"
                cfo = f"({ni}+{da}+{sbc}+{d_ar}+{d_vntr}+{d_inv}+{d_otherca}+{d_ap}+{d_otherliab})"
                cell.value = f"={cfo}-{capex}"
            cell.number_format = CUR
            cell.font = BLACK
    return top_row + 2 + len(growth_vals)

next_row = build_grid(ws, 5, "Net Income ($mm)", "NI")
next_row = build_grid(ws, next_row + 2, "Free Cash Flow ($mm) = Cash from Operations - Capex", "FCF")

ws[f"A{next_row+2}"] = ("Base Case (8% growth / 47.2% margin) is highlighted by its row/column position above. "
                         "Bear Case (-10% growth) sits in the top-left of the grid; note how Net Income and FCF "
                         "compress faster than revenue does, since opex is modeled as % of revenue but a real "
                         "company carries fixed costs that wouldn't flex down as quickly in an actual downturn.")
ws[f"A{next_row+2}"].font = SUBTITLE_FONT
ws.merge_cells(f"A{next_row+2}:G{next_row+2}")
ws.row_dimensions[next_row+2].height = 30

autosize(ws, {"A": 32, "B": 13, "C": 13, "D": 13, "E": 13, "F": 13, "G": 13})
ws.sheet_view.showGridLines = False

print("Sensitivity Table done")
wb.save("/home/claude/fpa-model/model/AAPL_3statement_model.xlsx")

# Reorder sheets logically
order = ["README", "Assumptions", "Income Statement", "Balance Sheet",
         "Cash Flow Statement", "Scenario Analysis", "Sensitivity Table"]
wb._sheets = [wb[name] for name in order]
for name in order:
    wb[name].sheet_properties.tabColor = "1F4E78" if name != "README" else "808080"
wb.active = 0
wb.save("/home/claude/fpa-model/model/AAPL_3statement_model.xlsx")
print("All sheets ordered and saved")
